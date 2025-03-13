import subprocess
import glob
import os
from datetime import date
import openpyxl
import numpy as np
from openpyxl.utils import get_column_letter


def afinfo(file):
    command = ["afinfo", file]
    result = subprocess.run(command, capture_output=True, text=True)
    return result.stdout

def parsed_data(af_out):
    data = {}
    i = 0
    lines = af_out.split("\n")

    while i < len(lines):
        line = lines[i].strip()
        if not line:
            i += 1
            continue

        if line.startswith("File:"):
            filepath = line.split(":", 1)[1].strip()
            data['File Path'] = filepath
            data["Artist"] = os.path.basename(os.path.dirname(os.path.dirname(filepath)))
            data["File"] = os.path.basename(filepath).split('.')[0]
            data["Album/ Single/ EP"] = os.path.basename(os.path.dirname(filepath))

        if "File type ID:" in line:
            data["File Type"] = line.split(":", 1)[1].strip()

        if "Num Tracks:" in line:
            data["Track Number"] = line.split(":", 1)[1].strip()

        if "Data format:" in line:
            data["Data Format"] = line.split(":", 1)[1].strip()
            parts = data["Data Format"].split(",")
            if len(parts) > 1:
                data["Channel"] = parts[0].strip()
                data["Sample Rate"] = parts[1].strip()

        if "estimated duration:" in line:
            number = line.split(":", 1)[1].strip()
            try:
                data["Duration (Min)"] = float(number.split()[0]) / 60
            except ValueError:
                data["Duration (Min)"] = None

        if "audio bytes:" in line:
            data["Bytes"] = line.split(":", 1)[1].strip()

        if "audio packets:" in line:
            data["Audio Packets"] = line.split(":", 1)[1].strip()

        if "bit rate:" in line:
            data["Bit Rate"] = line.split(":", 1)[1].strip()

        if "packet size upper bound:" in line:
            data["Packet Size Upper Bound"] = line.split(":", 1)[1].strip()

        if "maximum packet size:" in line:
            data["Maximum Packet Size"] = line.split(":", 1)[1].strip()

        if "audio data file offset:" in line:
            data["Audio Data File Offset"] = line.split(":", 1)[1].strip()

        if "source bit depth:" in line:
            bit_depth = line.split(":", 1)[1].strip()
            bit_depth_map = {
                "I16": "16 bit",
                "I8": "8 bit",
                "I24": "24 bit",
                "I32": "32 bit",
                "F32": "32-bit float",
                "F64": "64-bit float"
            }
            data["Source Bit Depth"] = bit_depth_map.get(bit_depth, "Unknown bit depth")

        if "Channel layout:" in line:
            data["Channel Layout"] = line.split(":", 1)[1].strip()

        if "Loudness Info:" in line:
            Loudness_Info = {}
            i += 1  # Move to next line

            while i < len(lines):
                line = lines[i].strip()

                if not line:
                    i += 1
                    continue 

                if "sound check volume normalization gain:" in line.lower():
                    break 

                if "album loudness parameters" in line.lower():
                    current_section = "Album Loudness Parameters"
                    Loudness_Info[current_section] = {}

                if "main loudness parameters" in line.lower():
                    current_section = "Main Loudness Parameters"
                    Loudness_Info[current_section] = {}

                if "sound check info" in line.lower():
                    current_section = "Sound Check Info"
                    Loudness_Info[current_section] = {}

                else:
                    key_value = line.split(":", 1)
                    if len(key_value) == 2 and key_value[1].strip():
                        key = key_value[0].strip()
                        value = key_value[1].strip()
                        Loudness_Info[current_section][key] = value

                i += 1

            data["Loudness Info"] = Loudness_Info

        if "sound check volume normalization gain:" in line:
            i += 1
            data["Sound Check Volume Normalization Gain"] = line.split(":", 1)[1].strip()

        i += 1

    return data

def Excel_Maker(metadata):
    day = date.today()
    base_dir = os.path.basename(os.path.dirname(os.path.dirname(os.path.dirname(metadata[0]['File Path']))))
    
    directory = f"Analysis/{base_dir}/{day}"

    current_path = ""
    for part in directory.split(os.sep):
        current_path = os.path.join(current_path, part)
        if not os.path.exists(current_path):
            os.mkdir(current_path)

    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = f'{base_dir}'


    file_path = f"{directory}/{base_dir}_{day}.xlsx"
    
    #General Info
    artist_name = []
    song_name   = []
    album       = []
    track_number= []
    duration    = []
    sample_rate = []
    bit_depth   = []

    #Loudness Parameters
    #Album Loudness Parameter
    album_aa_ebu_max_momentary_loudness  = []
    album_aa_ebu_top_of_loudness_range   = []
    album_aa_itu_sample_peak             = []
    album_aa_ebu_max_short_term_loudness = []
    album_aa_itu_true_peak               = []
    album_aa_ebu_loudness_range          = []
    album_aa_itu_loudness                = []

    #Main Loudness Parameter
    main_aa_ebu_max_momentary_loudness  = []
    main_aa_ebu_top_of_loudness_range   = []
    main_aa_itu_sample_peak             = []
    main_aa_itu_true_peak               = []
    main_aa_ebu_max_short_term_loudness = []
    main_aa_ebu_loudness_range          = []
    main_aa_itu_loudness                = []

    #Sound Check Info
    sc_ave_perceived_power_coeff   = []
    sc_max_perceived_power_coeff   = []
    sc_peak_amplitude_msec         = []
    sc_max_perceived_power_msec    = []
    sc_peak_amplitude              = []

    sound_Check_volume_normalization_gain = []


    for j in range(len(metadata)):
        
        print("File path: ", metadata[j]['File Path'], end='\n')

        artist_name.append(metadata[j]['Artist'])
        song_name.append(metadata[j]['File'])
        album.append(metadata[j]['Album/ Single/ EP'])
        track_number.append(metadata[j]['Track Number'])
        duration.append(metadata[j]['Duration (Min)'])
        sample_rate.append(metadata[j]['Sample Rate'])
        bit_depth.append(metadata[j]['Source Bit Depth'])

        try:
            main_aa_ebu_max_momentary_loudness.append(metadata[j]['Loudness Info']['Main Loudness Parameters']['aa ebu max momentary loudness'])
            main_aa_ebu_top_of_loudness_range.append(metadata[j]['Loudness Info']['Main Loudness Parameters']['aa ebu top of loudness range'])
            main_aa_itu_sample_peak.append(metadata[j]['Loudness Info']['Main Loudness Parameters']['aa itu sample peak'])
            main_aa_itu_true_peak.append(metadata[j]['Loudness Info']['Main Loudness Parameters']['aa itu true peak'])
            main_aa_ebu_max_short_term_loudness.append(metadata[j]['Loudness Info']['Main Loudness Parameters']['aa ebu max short-term loudness'])
            main_aa_ebu_loudness_range.append(metadata[j]['Loudness Info']['Main Loudness Parameters']['aa ebu loudness range'])
            main_aa_itu_loudness.append(metadata[j]['Loudness Info']['Main Loudness Parameters']['aa itu loudness'])
        except KeyError:

            main_aa_ebu_max_momentary_loudness.append("")
            main_aa_ebu_top_of_loudness_range.append("")
            main_aa_itu_sample_peak.append("")
            main_aa_itu_true_peak.append("")
            main_aa_ebu_max_short_term_loudness.append("")
            main_aa_ebu_loudness_range.append("")
            main_aa_itu_loudness.append("")

        try:
            album_aa_ebu_max_momentary_loudness.append(metadata[j]['Loudness Info']['Album Loudness Parameters'].get('aa ebu max momentary loudness', ""))
            album_aa_ebu_top_of_loudness_range.append(metadata[j]['Loudness Info']['Album Loudness Parameters'].get('aa ebu top of loudness range', ""))
            album_aa_itu_sample_peak.append(metadata[j]['Loudness Info']['Album Loudness Parameters'].get('aa itu sample peak', ""))
            album_aa_itu_true_peak.append(metadata[j]['Loudness Info']['Album Loudness Parameters'].get('aa itu true peak', ""))
            album_aa_ebu_max_short_term_loudness.append(metadata[j]['Loudness Info']['Album Loudness Parameters'].get('aa ebu max short-term loudness', ""))
            album_aa_ebu_loudness_range.append(metadata[j]['Loudness Info']['Album Loudness Parameters'].get('aa ebu loudness range', ""))
            album_aa_itu_loudness.append(metadata[j]['Loudness Info']['Album Loudness Parameters'].get('aa itu loudness', ""))

        except KeyError:
            album_aa_ebu_max_momentary_loudness.append("")
            album_aa_ebu_top_of_loudness_range.append("")
            album_aa_itu_sample_peak.append("")
            album_aa_itu_true_peak.append("")
            album_aa_ebu_max_short_term_loudness.append("")
            album_aa_ebu_loudness_range.append("")
            album_aa_itu_loudness.append("")
        try:
            sc_ave_perceived_power_coeff.append(metadata[j]['Loudness Info']['Sound Check Info']['sc ave perceived power coeff'])
            sc_max_perceived_power_coeff.append(metadata[j]['Loudness Info']['Sound Check Info']['sc max perceived power coeff'])
            sc_peak_amplitude_msec.append(metadata[j]['Loudness Info']['Sound Check Info']['sc peak amplitude msec'])
            sc_max_perceived_power_msec.append(metadata[j]['Loudness Info']['Sound Check Info']['sc max perceived power msec'])
            sc_peak_amplitude.append(metadata[j]['Loudness Info']['Sound Check Info']['sc peak amplitude'])

        except KeyError:
            sc_ave_perceived_power_coeff.append("")
            sc_max_perceived_power_coeff.append("")
            sc_peak_amplitude_msec.append("")
            sc_max_perceived_power_msec.append("")
            sc_peak_amplitude.append("")
        
        try:
            sound_Check_volume_normalization_gain.append(metadata[j]['Sound Check Volume Normalization Gain'])

        except KeyError:
            sound_Check_volume_normalization_gain.append("")

    headers = [
        "Sl.NO",
        "Artist",
        "Track",
        "Album/ Single/ EP",
        "Duration (Min)",
        "Sample Rate",
        "Source Bit Depth",

        # Main Loudness Parameters
        "Main AA EBU Max Momentary Loudness",
        "Main AA EBU Top Of Loudness Range",
        "Main AA ITU Sample Peak",
        "Main AA ITU True Peak",
        "Main AA EBU Max Short Term Loudness",
        "Main AA EBU Loudness Range",
        "Main AA ITU Loudness",

        # Album Loudness Parameters
        "Album AA EBU Max Momentary Loudness",
        "Album AA EBU Top Of Loudness Range",
        "Album AA ITU Sample Peak",
        "Album AA ITU True Peak",
        "Album AA EBU Max Short Term Loudness",
        "Album AA EBU Loudness Range",
        "Album AA ITU Loudness",


        # Sound Check Info
        "SC Ave Perceived Power Coeff",
        "SC Max Perceived Power Coeff",
        "SC Peak Amplitude Msec",
        "SC Max Perceived Power Msec",
        "SC Peak Amplitude",
        "Sound Check Volume Normalization Gain"
        ]

    sheet.append(headers)
    
    sl_no = np.arange(1, len(artist_name) + 1)

    for i in range(len(artist_name)):
        sheet.cell(row=i + 2, column=1, value=sl_no[i])
        sheet.cell(row=i + 2, column=2, value=artist_name[i])
        sheet.cell(row=i + 2, column=3, value=song_name[i])
        sheet.cell(row=i + 2, column=4, value=album[i])
        sheet.cell(row=i + 2, column=5, value=str(duration[i]))
        sheet.cell(row=i + 2, column=6, value=sample_rate[i])
        sheet.cell(row=i + 2, column=7, value=bit_depth[i])

        sheet.cell(row=i + 2, column=8, value=main_aa_ebu_max_momentary_loudness[i])
        sheet.cell(row=i + 2, column=9, value=main_aa_ebu_top_of_loudness_range[i])
        sheet.cell(row=i + 2, column=10, value=main_aa_itu_sample_peak[i])
        sheet.cell(row=i + 2, column=11, value=main_aa_itu_true_peak[i])
        sheet.cell(row=i + 2, column=12, value=main_aa_ebu_max_short_term_loudness[i])
        sheet.cell(row=i + 2, column=13, value=main_aa_ebu_loudness_range[i])
        sheet.cell(row=i + 2, column=14, value=main_aa_itu_loudness[i])

        sheet.cell(row=i + 2, column=15, value=album_aa_ebu_max_momentary_loudness[i])
        sheet.cell(row=i + 2, column=16, value=album_aa_ebu_top_of_loudness_range[i])
        sheet.cell(row=i + 2, column=17, value=album_aa_itu_sample_peak[i])
        sheet.cell(row=i + 2, column=18, value=album_aa_itu_true_peak[i])
        sheet.cell(row=i + 2, column=19, value=album_aa_ebu_max_short_term_loudness[i])
        sheet.cell(row=i + 2, column=20, value=album_aa_ebu_loudness_range[i])
        sheet.cell(row=i + 2, column=21, value=album_aa_itu_loudness[i])

        sheet.cell(row=i + 2, column=22, value=sc_ave_perceived_power_coeff[i])
        sheet.cell(row=i + 2, column=23, value=sc_max_perceived_power_coeff[i])
        sheet.cell(row=i + 2, column=24, value=sc_peak_amplitude_msec[i])
        sheet.cell(row=i + 2, column=25, value=sc_max_perceived_power_msec[i])
        sheet.cell(row=i + 2, column=26, value=sc_peak_amplitude[i])
        sheet.cell(row=i + 2, column=27, value=sound_Check_volume_normalization_gain[i])

        


        
        



    sheet.row_dimensions[1].height = 48
    sheet.column_dimensions['B'].width = 35
    sheet.column_dimensions['C'].width = 40
    sheet.column_dimensions['D'].width = 66 
    for a in range(5, len(headers)+1):
        column_letter = get_column_letter(a)
        sheet.column_dimensions[column_letter].width = 40
    
    wb.save(file_path)
    print(f"{os.path.basename(file_path)} is saved in location {os.path.dirname(file_path)}")
        

music_dir = input("Enter Music directory path: ")
file_paths = glob.glob(f"{music_dir}/**/*.m4p", recursive=True)


dataframe = []
for file in file_paths:
    af_out = afinfo(file)
    data = parsed_data(af_out)
    dataframe.append(data)

Excel_Maker(dataframe)